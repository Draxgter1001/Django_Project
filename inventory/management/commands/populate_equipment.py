from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from inventory.models import Equipment, Location
from django.db import transaction


class Command(BaseCommand):
    help = 'Load equipment data from Excel file using openpyxl'

    def handle(self, *args, **options):
        file_path = r'/Users/uthsh/PycharmProjects/djangoProjectGroup2/media/equipment.xlsx'
        wb = load_workbook(filename=file_path)
        ws = wb['MainInventory']

        with transaction.atomic():
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 7:  # Ensuring the row exists and has enough columns
                    continue  # Skip incomplete rows

                # Using a default location name if none is provided
                location_name = row[4] if row[4] else "Default Location"

                location, _ = Location.objects.get_or_create(
                    location_name=location_name,
                    defaults={'location_type': 'Default'}
                )

                # Proceed to create/update Equipment only if a name is provided
                if row[0]:  # Checking if equipment name exists
                    Equipment.objects.update_or_create(
                        name=row[0],
                        defaults={
                            'type': row[1],
                            'quantity': row[2],
                            'last_audit': row[3],
                            'location': location,
                            'status': row[5] if row[5] else '',
                            'comments': row[6] if row[6] else ''
                        }
                    )
                else:
                    print(f"Skipping equipment creation due to missing name for row: {row}")

        print('Data import complete')
